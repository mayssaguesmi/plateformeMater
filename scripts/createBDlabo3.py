import mysql.connector
from slugify import slugify
import hashlib
import sys
from openpyxl import load_workbook

# -------------------------
# CONFIG
# -------------------------
DB_CONFIG = {
    "host": "localhost",
    "user": "phpmyadmin",
    "password": "admin",
    "database": "utmsearch_w"
}

EXCEL_FILE = "Liste des Labos de recherche UTM.xlsx"

# -------------------------
# Connexion DB
# -------------------------
try:
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor(dictionary=True)
    print("✅ Connexion DB réussie")
except Exception as e:
    print(f"❌ Erreur connexion DB : {e}")
    sys.exit(1)

# -------------------------
# Fonctions utilitaires
# -------------------------
def get_etablissement_id(nom):
    try:
        if not nom:
            return None
        nom_clean = str(nom).strip().lower()
        cursor.execute("SELECT id, nom FROM utm_master_instituts")
        rows = cursor.fetchall()
        for r in rows:
            if (r["nom"].lower() == nom_clean or
                r["nom"].upper() == nom.upper() or
                r["nom"].capitalize() == nom.capitalize()):
                return r["id"]
    except Exception as e:
        print(f"⚠️ Erreur recherche établissement '{nom}': {e}")
    return None

def get_or_create_directeur(nom_responsable, email, etablissement_id):
    try:
        if not email:
            return None
        
        cursor.execute("SELECT ID FROM utm_users WHERE user_email=%s", (email,))
        row = cursor.fetchone()
        if row:
            return row["ID"]

        login = slugify(email.split("@")[0])
        pwd = hashlib.md5("changeme123".encode()).hexdigest()
        cursor.execute("""
            INSERT INTO utm_users (user_login, user_pass, user_nicename, user_email, display_name)
            VALUES (%s, %s, %s, %s, %s)
        """, (login, pwd, nom_responsable or login, email, nom_responsable or login))
        conn.commit()
        user_id = cursor.lastrowid

        cursor.execute("""
            INSERT INTO utm_usermeta (user_id, meta_key, meta_value) VALUES (%s, %s, %s)
        """, (user_id, "utm_capabilities", '{"um_directeur_laboratoire":true}'))

        if etablissement_id:
            cursor.execute("""
                INSERT INTO utm_usermeta (user_id, meta_key, meta_value) VALUES (%s, %s, %s)
            """, (user_id, "institut_id", str(etablissement_id)))

        conn.commit()
        print(f"👤 Utilisateur créé : {nom_responsable} ({email}) [ID={user_id}]")
        return user_id
    except Exception as e:
        print(f"⚠️ Erreur création user {email}: {e}")
        return None

def laboratoire_exists(code_lr, denomination):
    try:
        if code_lr:
            cursor.execute("SELECT id FROM utm_recherche_laboratoire WHERE code_lr=%s", (code_lr,))
            if cursor.fetchone():
                return True
        if denomination:
            cursor.execute("SELECT id FROM utm_recherche_laboratoire WHERE denomination=%s", (denomination,))
            if cursor.fetchone():
                return True
    except Exception as e:
        print(f"⚠️ Erreur check doublon labo : {e}")
    return False

# -------------------------
# Lecture Excel avec indices
# -------------------------
try:
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    print(f"📄 Fichier Excel chargé : {EXCEL_FILE}")
except Exception as e:
    print(f"❌ Erreur ouverture Excel : {e}")
    sys.exit(1)

# Définir les indices des colonnes (0 = première colonne)
IDX_NUMERO_ETAB     = 1   # N°/Etab
IDX_SR              = 3   # SR
IDX_DISCIPLINE      = 4   # Discipline
IDX_DENOMINATION    = 5   # Dénomination
IDX_DENOMINATION_AR = 6   # التسمية
IDX_CODE_LR         = 7   # Code
IDX_DIRECTEUR_NOM   = 8   # Nom du responsable
IDX_EMAIL           = 9   # Email
IDX_ETAB            = 10  # Etablissement
IDX_SITE_WEB        = 11  # Site web

# -------------------------
# Parcours et insertion
# -------------------------
for row in ws.iter_rows(min_row=7, values_only=True):  # ⚠️ adapter min_row si l'entête finit à la ligne 6
    numero_etab     = row[IDX_NUMERO_ETAB]
    sr              = row[IDX_SR]
    discipline      = row[IDX_DISCIPLINE]
    denomination    = row[IDX_DENOMINATION]
    denomination_ar = row[IDX_DENOMINATION_AR]
    code_lr         = row[IDX_CODE_LR]
    directeur_nom   = row[IDX_DIRECTEUR_NOM]
    directeur_email = row[IDX_EMAIL]
    etablissement_label = row[IDX_ETAB]
    site_web        = row[IDX_SITE_WEB]

    # Ignorer si infos essentielles manquantes
    if not denomination or not code_lr:
        continue

    
    # Vérifier doublons
    if laboratoire_exists(code_lr, denomination):
        print(f"⚠️ Labo déjà existant : {denomination} ({code_lr}) → ignoré")
        continue

    etablissement_id = get_etablissement_id(etablissement_label)


    directeur_user_id = get_or_create_directeur(directeur_nom, directeur_email, etablissement_id)

    try:
        cursor.execute("""
            INSERT INTO utm_recherche_laboratoire 
            (numero_etab, sr, domaine, denomination, denomination_ar, code_lr,
             directeur_nom, directeur_email, directeur_user_id,
             etablissement_id, etablissement_label, site_web, statut)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'Actif')
        """, (
            numero_etab,
            sr,
            discipline,
            denomination,
            denomination_ar,
            code_lr,
            directeur_nom,
            directeur_email,
            directeur_user_id,
            etablissement_id,
            etablissement_label,
            site_web
        ))
        conn.commit()
        print(f"✅ Labo inséré : {denomination} ({code_lr})")

    except Exception as e:
        print(f"❌ Erreur insertion labo {denomination}: {e}")
        conn.rollback()

print("🚀 Import terminé avec lecture par index")
